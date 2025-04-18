from django.contrib import admin
from .models import (
    Vehiculo, ChequeoDiario, MantenimientoVehiculo,
    DocumentoVehiculo, AsignacionVehiculo,
    Oficial, ArticuloInventario, MovimientoInventario,
    Arma, MovimientoArma, UniformeAsignado, CargaCombustible
)
from django.http import HttpResponse
from django.db import models
from django.db.models import Sum
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from rangefilter.filters import DateRangeFilter

# --- Funciones de exportación ---
def exportar_a_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={modeladmin.model.__name__}_export.csv'
    writer = csv.writer(response)
    writer.writerow([field.name for field in modeladmin.model._meta.fields])
    for obj in queryset:
        writer.writerow([getattr(obj, field.name) for field in modeladmin.model._meta.fields])
    return response
exportar_a_csv.short_description = "Exportar a CSV"

def exportar_a_excel(modeladmin, request, queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = modeladmin.model.__name__
    fields = [field.name for field in modeladmin.model._meta.fields]

    for col_num, column_title in enumerate(fields, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = column_title

    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(fields, 1):
            value = getattr(obj, field_name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={modeladmin.model.__name__}_export.xlsx'
    workbook.save(response)
    return response
exportar_a_excel.short_description = "Exportar a Excel (.xlsx)"

# Exportar cargas con bajo rendimiento
def exportar_bajo_rendimiento(modeladmin, request, queryset):
    bajo_rendimiento = queryset.filter(
        kilometraje_final__isnull=False,
        galones_cargados__gt=0
    ).annotate(
        rendimiento=models.ExpressionWrapper(
            (models.F('kilometraje_final') - models.F('kilometraje_inicial')) / models.F('galones_cargados'),
            output_field=models.FloatField()
        )
    ).filter(rendimiento__lt=6)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bajo Rendimiento"

    columnas = ["Vehículo", "Fecha", "Km Inicial", "Km Final", "Km Recorridos", "Galones", "Rendimiento", "Costo Total"]
    sheet.append(columnas)

    for obj in bajo_rendimiento:
        recorrido = obj.kilometraje_final - obj.kilometraje_inicial
        rendimiento = recorrido / obj.galones_cargados if obj.galones_cargados else 0
        costo = obj.galones_cargados * obj.precio_por_galon
        sheet.append([
            obj.vehiculo.placa,
            obj.fecha,
            obj.kilometraje_inicial,
            obj.kilometraje_final,
            recorrido,
            obj.galones_cargados,
            round(rendimiento, 2),
            f"RD${round(costo, 2)}"
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Cargas_Bajo_Rendimiento.xlsx'
    workbook.save(response)
    return response
exportar_bajo_rendimiento.short_description = "📉 Exportar Bajo Rendimiento (<6 Km/Gal)"

# --- Admins ---
@admin.register(Vehiculo)
class VehiculoAdmin(admin.ModelAdmin):
    list_display = ['placa', 'marca', 'modelo', 'anio', 'tipo', 'departamento', 'kilometraje']
    search_fields = ['placa', 'marca', 'modelo']
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(ChequeoDiario)
class ChequeoDiarioAdmin(admin.ModelAdmin):
    list_display = ['vehiculo', 'fecha', 'observaciones']
    list_filter = [('fecha', DateRangeFilter)]
    search_fields = ['vehiculo__placa']
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(MantenimientoVehiculo)
class MantenimientoVehiculoAdmin(admin.ModelAdmin):
    list_display = ['vehiculo', 'fecha_mantenimiento', 'tipo_mantenimiento', 'kilometraje']
    list_filter = [('fecha_mantenimiento', DateRangeFilter)]
    search_fields = ['vehiculo__placa', 'tipo_mantenimiento']
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(DocumentoVehiculo)
class DocumentoVehiculoAdmin(admin.ModelAdmin):
    list_display = ['vehiculo', 'tipo', 'fecha_vencimiento']
    list_filter = ['tipo']
    search_fields = ['vehiculo__placa']
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(AsignacionVehiculo)
class AsignacionVehiculoAdmin(admin.ModelAdmin):
    list_display = ['vehiculo', 'supervisor', 'zona', 'fecha_asignacion', 'fecha_vencimiento_licencia']
    list_filter = ['zona']
    search_fields = ['vehiculo__placa', 'supervisor__nombre']
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(CargaCombustible)
class CargaCombustibleAdmin(admin.ModelAdmin):
    list_display = [
        'vehiculo', 'fecha', 'kilometraje_inicial', 'kilometraje_final',
        'galones_cargados', 'precio_por_galon', 'mostrar_rendimiento_alerta'
    ]
    list_filter = [('fecha', DateRangeFilter)]
    search_fields = ['vehiculo__placa']
    actions = [exportar_a_csv, exportar_a_excel, exportar_bajo_rendimiento]

    def mostrar_rendimiento_alerta(self, obj):
        if obj.kilometraje_final and obj.galones_cargados:
            recorrido = obj.kilometraje_final - obj.kilometraje_inicial
            if obj.galones_cargados > 0:
                rendimiento = recorrido / obj.galones_cargados
                icono = "✅" if rendimiento >= 6 else "🚨"
                return f"{icono} {round(rendimiento, 2)}"
        return "-"
    mostrar_rendimiento_alerta.short_description = "Rend. Km/Gal"



@admin.register(Oficial)
class OficialAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'cedula', 'cargo', 'ubicacion', 'estado', 'fecha_ingreso')
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(Arma)
class ArmaAdmin(admin.ModelAdmin):
    list_display = ['codigo_arma', 'serial', 'tipo', 'estado', 'almacen_actual', 'asignada_a', 'fija_en_puesto', 'fecha_vencimiento_licencia', 'vence_pronto_display']
    list_filter = ['tipo', 'estado', 'transferida']
    search_fields = ['serial']
    actions = [exportar_a_csv, exportar_a_excel]

    def vence_pronto_display(self, obj):
        return "✅ Sí" if obj.vence_pronto else "❌ No"
    vence_pronto_display.short_description = "¿Vence pronto?"

@admin.register(ArticuloInventario)
class ArticuloInventarioAdmin(admin.ModelAdmin):
    list_display = ['nombre', 'categoria', 'cantidad_total', 'cantidad_disponible', 'unidad', 'talla', 'serial']
    list_filter = ['categoria']
    search_fields = ['nombre', 'descripcion', 'serial']
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(MovimientoInventario)
class MovimientoInventarioAdmin(admin.ModelAdmin):
    list_display = ['articulo', 'tipo', 'cantidad', 'fecha', 'responsable']
    list_filter = [('fecha', DateRangeFilter), 'tipo']
    search_fields = ['articulo__nombre', 'responsable']
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(MovimientoArma)
class MovimientoArmaAdmin(admin.ModelAdmin):
    list_display = ['arma', 'tipo', 'oficial', 'fecha']
    list_filter = [('fecha', DateRangeFilter), 'tipo']
    search_fields = ['arma__serial', 'oficial__nombre']
    actions = [exportar_a_csv, exportar_a_excel]

@admin.register(UniformeAsignado)
class UniformeAsignadoAdmin(admin.ModelAdmin):
    list_display = ['oficial', 'articulo', 'talla', 'estado', 'fecha_entrega']
    list_filter = ['estado', 'fecha_entrega']
    search_fields = ['oficial__nombre', 'articulo__nombre', 'talla']
    actions = [exportar_a_csv, exportar_a_excel]
